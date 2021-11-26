from numpy.core.fromnumeric import shape
import tensorflow as tf
import keras 
from keras.models import Sequential, Input, Model
from keras.layers import Dense, Dropout, Flatten, Conv1D
from keras.layers import Conv1D, MaxPooling1D
#from keras.layers.normalization import Batchnormalization
from keras.layers.advanced_activations import LeakyReLU, ReLU, Softmax
import pandas as pd 
import numpy as np
import openpyxl
from openpyxl import load_workbook
import pickle



#Twoflo_data = pd.read_excel(r'C:\Users\wardv\Documents\Ward\KUL\MIIµ\Thesis\Sampled\Sampled_Recordings.xlsx','Sheet1',header=None)
#Triflo_data = pd.read_excel(r'C:\Users\wardv\Documents\Ward\KUL\MIIµ\Thesis\Sampled\Sampled_Recordings.xlsx','Sheet2',header=None)

#shape = Triflo_data.shape
#print(shape)

#Array_1 = Triflo_data

workbook = load_workbook(r'C:\Users\wardv\Documents\Ward\KUL\MIIµ\Thesis\Sampled\Sampled_Recordings.xlsx')

sheets = workbook.sheetnames
print(sheets)

Sheet1 = workbook['TwoFlo']
Sheet2 = workbook['TriFlo']

row_max = Sheet1.max_row
col_max = Sheet1.max_column

TwoFlo_data = []



for i in range(1,col_max+1):
    current_column = []
    for j in range(1,row_max+1):
        current_column.append(Sheet1.cell(j,i).value)
    TwoFlo_data.append(current_column)

row_max = Sheet2.max_row
col_max = Sheet2.max_column

TwoFlo_data = np.array(TwoFlo_data)

print(shape(TwoFlo_data))
print(TwoFlo_data)

TwoFlo_training = TwoFlo_data[0:1999]
TwoFlo_validation = TwoFlo_data[2000:2416]

TriFlo_data = []

for i in range(1,col_max+1):
    current_column = []
    for j in range(1,row_max+1):
        current_column.append(Sheet2.cell(j,i).value)
    TriFlo_data.append(current_column)

TriFlo_data = np.array(TriFlo_data)

print(shape(TriFlo_data))
TriFlo_training = TriFlo_data[0:899]
print(TriFlo_training)
Triflo_validation = TriFlo_data[900:1155]

print(TwoFlo_training)
print(TriFlo_training)

Train_X = np.vstack((TwoFlo_training,TriFlo_training))
Validate_X = np.vstack((TwoFlo_validation,Triflo_validation))

Train_1 = np.zeros(1999)

Train_2 = np.ones(899)



Train_Y = np.hstack((Train_1,Train_2))
print(Train_Y)

Validate_1 = np.zeros(416)

Validate_2 = np.ones(255)

Validate_Y = np.hstack((Validate_1,Validate_2))

pickle.dump(Train_X,open("trainx.p","wb"))
pickle.dump(Train_Y,open("trainy.p","wb"))
pickle.dump(Validate_X,open("validatex.p","wb"))
pickle.dump(Validate_Y,open("validatey.p","wb"))



